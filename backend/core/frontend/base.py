import io
import logging
from typing import Optional, Any, Callable
from uuid import UUID

import openpyxl
from fastapi import APIRouter, UploadFile, File
from fastapi import Request
from fastapi.params import Depends
from fastapi.responses import HTMLResponse
from pydantic import BaseModel
from starlette.responses import JSONResponse

from app.front.front_tasks import import_prepare_data, import_save
from core.frontend.constructor import ClassView, Method, views, get_model
from core.frontend.utils import clean_filter

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ExceptionResponseSchema(BaseModel):
    error: str


router = APIRouter(
    responses={"400": {"model": ExceptionResponseSchema}},
)


@router.get("/filter/{model}", response_class=HTMLResponse)
async def _filter(request: Request, model: str):
    """
     Универсальный запрос, который отдает фильтр обьекта по его модулю и модели
    """
    cls_view = views.get(model)
    if not cls_view:
        cls = ClassView(request, model)
    else:
        cls = cls_view(request)
    return await cls.h.as_filter_get


@router.get("/search", response_class=JSONResponse)
async def search(request: Request, model: str, q: str, filter: str):
    """
     Универсальный запрос поиска
    """
    params = {'search': q}
    if f := filter:
        if isinstance(f, str):
            try:
                filter = eval(f)
            except TypeError as ex:
                raise Exception(f'Filter {f} is not valid') from ex
    if isinstance(filter, dict):
        params.update(filter)
    async with request.scope['env'][model].adapter as a:
        data = await a.list(params=params, model=model)
    return [
        {
            'value': i['id'],
            'label': i.get('title') or i.get('name') or i.get('english_name') or i.get('nickname')
        }
        for i in data['data']
    ]


@router.get("/get_by_ids", response_class=JSONResponse)
async def get_by_ids(request: Request, model: str, id__in: str):
    """
     Универсальный запрос поиска
    """
    params = {'id__in': id__in}
    async with request.scope['env'][model].adapter as a:
        data = await a.list(params=params, model=model)
    return [
        {
            'value': i['id'],
            'label': i.get('title') or i.get('name') or i.get('english_name') or i.get('nickname')
        }
        for i in data['data']
    ]


@router.get("/table/{model}", response_class=HTMLResponse)
async def table(request: Request, model: str, cls: ClassView = Depends(get_model)):
    """
     Универсальный запрос, который отдает таблицу обьекта и связанные если нужно
    """
    await cls.init(params=dict(request.query_params))
    return await cls.h.as_table_get


@router.get("/line/{model}", response_class=HTMLResponse)
async def line(request: Request, model: str, parent_field: str = None):
    """
     Универсальный запрос, который отдает/изменяет обьект
    """
    cls_view = views.get(model)
    if not cls_view:
        cls = ClassView(request, model, parent_field=parent_field)
    else:
        cls = cls_view(request)
        cls.v.parent_field = parent_field
    """Отдать обьект на создание, в зависимости от mode (tr/div)"""
    return await cls.h.as_tr_create


@router.get("/line/{model}/{line_id}", response_class=HTMLResponse)  # type: ignore
async def line(request: Request, model: str, line_id: UUID, mode: str = 'tr', method: str = 'get'):
    """
     Универсальный запрос, который отдает/изменяет обьект
    """
    cls_view = views.get(model)
    if not cls_view:
        cls = ClassView(request, model)
    else:
        cls = cls_view(request)
    match method:
        case Method.UPDATE:
            """Отдать обьект на редактирование, в зависимости от mode (tr/div)"""
            line = await cls.get_lines(ids=[line_id], join_related=False)
            return await getattr(line.h, f'as_{mode}_update')
        case Method.GET:
            """Отдать обьект на чтение, в зависимости от mode (tr/div)"""
            line = await cls.get_lines(ids=[line_id], join_related=False)
            return await getattr(line.h, f'as_{mode}_get')
        case Method.CREATE:
            """Отдать обьект на создание, в зависимости от mode (tr/div)"""
            return await cls.h.as_tr_create
        case Method.DELETE:
            """Отдать обьект на удаление, в не зависимости от mode (tr/div)"""
            if isinstance(cls.v.schema.id, int):
                """Если это временная запись, то просто удалить"""
                return
            line = await cls.get_lines(ids=[line_id], join_related=False)
            return await line.h.as_modal_delete
        case Method.UPDATE_SAVE:
            """Сохранение записи при измененнии"""
            data = clean_filter(cls.v.schema.model_extra, cls.v.schema.key)
            await cls.update_lines(id=line_id, data=data)
        case Method.CREATE_SAVE:
            """Сохранение записи при создании"""
            data = clean_filter(cls.v.schema.model_extra, cls.v.schema.key)
            line = await cls.create_lines(data)
            return await line.h.as_div_update
        case Method.DELETE_SAVE:
            await cls.delete_lines(ids=[line_id])
            """Отдать обьект на удаление, в не зависимости от mode (tr/div)"""


@router.post("/line/{model}", response_class=HTMLResponse)  # type: ignore
async def line(request: Request, model: str):
    """
     Универсальный запрос, который отдает/изменяет обьект
    """
    cls_view = views.get(model)
    if not cls_view:
        cls = ClassView(request, model)
    else:
        cls = cls_view(request)
    """Сохранение записи при создании"""
    data = await request.json()
    line = await cls.create_line(data)
    return await line.h.as_div_update


@router.delete("/line/{model}/{line_id}", response_class=HTMLResponse)
async def line_delete(request: Request, model: str, line_id: UUID):
    """
     Универсальный запрос, который отдает/изменяет обьект
    """
    cls_view = views.get(model)
    if not cls_view:
        cls = ClassView(request, model)
    else:
        cls = cls_view(request)
    await cls.delete_lines(ids=[line_id])
    """Отдать обьект на удаление, в не зависимости от mode (tr/div)"""


@router.put("/line/{model}/{line_id}", response_class=HTMLResponse)
async def line_update(request: Request, model: str, line_id: UUID):
    """
     Универсальный запрос, который отдает/изменяет обьект
    """
    cls_view = views.get(model)
    if not cls_view:
        cls = ClassView(request, model)
    else:
        cls = cls_view(request)
    """Сохранение записи при измененнии"""
    data = await request.json()
    await cls.update_line(id=line_id, data=data)


@router.get("/modal/{model}/{line_id}", response_class=HTMLResponse)
async def modal(request: Request, model: str, line_id: Optional[str | UUID] = None, method: str = 'get'):
    """
     Универсальный запрос модалки, который отдает форму модели
    """
    cls_view = views.get(model)
    if not cls_view:
        cls = ClassView(request, model)
    else:
        cls = cls_view(request)
    match method:
        case Method.GET:
            line = await cls.get_lines(ids=[line_id])
            return await line.h.as_modal_get
        case Method.UPDATE:
            line = await cls.get_lines(ids=[line_id])
            return await line.h.as_modal_update
        case Method.DELETE:
            line = await cls.get_lines(ids=[line_id])
            return line.h.as_modal_delete
        case Method.CREATE:
            return await cls.h.as_modal_create


@router.get("/field/{model}/{line_id}/{field_name}", response_class=HTMLResponse)
async def field_get(request: Request, model: str, line_id: UUID, field_name: str):
    """
     Универсальный запрос на получение as_update поля
    """
    cls_view = views.get(model)
    if not cls_view:
        cls = ClassView(request, model)
    else:
        cls = cls_view(request)
    line = await cls.get_lines(ids=[line_id])
    return getattr(line, field_name).as_('update', button=True)


@router.put("/field/{model}/{line_id}/{field_name}", response_class=HTMLResponse)
async def field_update(request: Request, model: str, line_id: UUID, field_name: str):
    """
     Универсальный запрос на получение as_update поля
    """
    cls_view = views.get(model)
    if not cls_view:
        cls = ClassView(request, model)
    else:
        cls = cls_view(request)
    data = await request.json()
    async with cls.v.model.adapter as a:
        line = await a.get(id=line_id)
        line[field_name] = data[field_name]
        await cls.update_line(id=line_id, data=line)
    return


class ActionSchema(BaseModel):
    action: str
    ids: Optional[list[str]] = []
    _schema: Any = None
    commit: Optional[bool] = False

    class Config:
        extra = 'allow'


@router.get("/action/{model}/{action_name}/{line_id}", response_class=HTMLResponse)
async def action(request: Request, model: str, action_name: str, line_id: UUID):
    """
     Универсальный запрос, который отдает форму модели (черпает из ModelUpdateSchema
    """
    cls_view = views.get(model)
    if not cls_view:
        cls = ClassView(request, model)
    else:
        cls = cls_view(request)
    return await cls.get_action(action_name=action_name, line_id=line_id)


@router.post("/action/{model}/{action_name}/{line_id}", response_class=HTMLResponse)  # type: ignore
async def action(request: Request, model: str, action_name: str, line_id: UUID):
    """
     Универсальный запрос, который отдает форму модели (черпает из ModelUpdateSchema
    """
    cls_view = views.get(model)
    if not cls_view:
        cls = ClassView(request, model)
    else:
        cls = cls_view(request)
    func: Callable = getattr(cls.v.model.adapter, action_name)
    result = []
    if action_schema := cls.v.actions[action_name].get('schema'):
        data = await request.json()
        obj = action_schema(**data)
        res = await func(schema=obj)
        result += res
        return cls.send_message(message=f'Action {action_name} done')
    else:
        await func(payload={'id': line_id})
        return cls.send_message(message=f'Action {action_name} done')


class ImportSchema(BaseModel):
    class Config:
        extra = 'allow'


@router.get("/import/{model}", response_class=HTMLResponse)  # type: ignore
async def modal(request: Request, model: str, method: str):
    """
     Универсальный запрос модалки, который отдает форму модели
    """
    cls_view = views.get(model)
    if not cls_view:
        cls = ClassView(request, model)
    else:
        cls = cls_view(request)
    return cls.get_import


@router.put("/import/{model}", response_class=HTMLResponse)  # type: ignore
async def modal(request: Request, model: str, method: str):
    """
     Универсальный запрос модалки, который отдает форму модели
    """
    cls_view = views.get(model)
    if not cls_view:
        cls = ClassView(request, model)
    else:
        cls = cls_view(request)
    data = await request.json()
    lines = await import_save(model, data)
    await cls.init(params={}, data=lines)
    return cls.as_table


@router.post("/import_upload/model", response_class=HTMLResponse)
async def import_upload(request: Request, model: str, file: UploadFile = File(...)):
    """
     Универсальный запрос модалки, который отдает форму модели
    """
    model, key = request.query_params.values()
    cls_view = views.get(model)
    if not cls_view:
        cls = ClassView(request, model)
    else:
        cls = cls_view(request)
    format: str = file.filename.split('.')[-1]  # type: ignore
    data: list = []
    header: list = []
    import_schema: ClassView = cls.v.model.schemas.create
    match format:
        case 'csv':
            ...
        case 'xlsx':
            # Read it, 'f' type is bytes
            f = await file.read()
            xlsx = io.BytesIO(f)
            wb = openpyxl.load_workbook(xlsx)
            ws = wb[wb.sheetnames[0]]

            for i, cells in enumerate(ws.iter_rows()):
                if i == 0:
                    header = [cell.value.lower() for cell in cells]
                    continue
                line: dict = {}
                for col, label in enumerate(header):
                    line[label] = cells[col].value
                data.append(line)
    if not data:
        return f"{cls.send_message(message=f'No data in file')} {cls.get_import}"
    task = await import_prepare_data.kiq(model=cls.model_name, data=data)
    task_result = await task.wait_result()
    cls.errors = task_result.return_value[0]
    lines = task_result.return_value[1]
    if task_result.return_value[2] == 'update':
        import_schema = cls.model.schemas.update
    await cls.init(data=lines, schema=import_schema)
    return f"{cls.get_import_errors}\n{cls.as_table_update}"
